#!/usr/bin/env python
# pylint: disable=C0116,W0613
# This program is dedicated to the public domain under the CC0 license.

"""
Basic example for a bot that uses inline keyboards. For an in-depth explanation, check out
 https://git.io/JOmFw.
"""
import builtins
import logging

from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import Updater, CommandHandler, CallbackQueryHandler, CallbackContext

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s', level=logging.INFO
)
logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from openpyxl.descriptors.base import String

wb = load_workbook(filename = 'names.xlsx')
ws = wb.active
names = []
names_no_dup = []
family_name = ''
for row in ws.iter_rows(min_row=2, max_col=3):
   # names.append((row[0].value[::-1],row[1].value[::-1],row[2].value))
    names.append((row[0].value[::-1]))
s = list(set(names))
for n in s:
    names_no_dup.append([n])

for family in names_no_dup:
    for row in ws.iter_rows(min_row=2, max_col=3):
        if family[0] == row[0].value[::-1]:
            family.append([row[1].value[::-1],row[2].value])

is_personal_name = False


def findNumber(family_names,personal_names):
    print(family_names + ' ' + personal_names)
    for value in names_no_dup :
        if value[0] == family_names:
            for names in value:
                if names[0] == personal_names:
                    return names[1]
    return None


def start(update: Update, context: CallbackContext) -> None:
    global is_personal_name
    help_button_list = []
    button_list = []
    """Sends a message with three inline buttons attached."""
    for value in names_no_dup:
        help_button_list.append(InlineKeyboardButton(value[0][::-1],callback_data = value[0][::-1]))
    button_list.append(help_button_list)
    keyboard = button_list
    reply_markup = InlineKeyboardMarkup(keyboard)
    is_personal_name = True

    update.message.reply_text('Please choose a family name:', reply_markup=reply_markup)


def button(update: Update, context: CallbackContext) -> None:
    """Parses the CallbackQuery and updates the message text."""
    global is_personal_name
    global family_name
    help_button_list = []
    button_list = []
    query = update.callback_query
    # CallbackQueries need to be answered, even if no notification to the user is needed
    # Some clients may have trouble otherwise. See https://core.telegram.org/bots/api#callbackquery
    query.answer()
    if is_personal_name:
        family_name = query.data
        for value in names_no_dup:
            if value[0][::-1] == query.data:
                for name in value[1::]:
                  help_button_list.append(InlineKeyboardButton(name[0][::-1],callback_data = name[0][::-1]))
        button_list.append(help_button_list)
        keyboard = button_list
        reply_markup = InlineKeyboardMarkup(keyboard)
        query.message.reply_text('Please choose a personaly name:', reply_markup=reply_markup)
        is_personal_name = False
    else :
         query.edit_message_text(text=f"phone number: {findNumber(family_name[::-1],query.data[::-1])}")

        

    


def help_command(update: Update, context: CallbackContext) -> None:
    """Displays info on how to use the bot."""
    update.message.reply_text("Use /start to test this bot.")


def main() -> None:
    """Run the bot."""
    # Create the Updater and pass it your bot's token.
    updater = Updater(token=#YOUR_TOKEN, use_context=True)

    updater.dispatcher.add_handler(CommandHandler('start', start))
    updater.dispatcher.add_handler(CallbackQueryHandler(button))
    updater.dispatcher.add_handler(CommandHandler('help', help_command))

    # Start the Bot
    updater.start_polling()

    # Run the bot until the user presses Ctrl-C or the process receives SIGINT,
    # SIGTERM or SIGABRT
    updater.idle()


if __name__ == '__main__':
    main()
