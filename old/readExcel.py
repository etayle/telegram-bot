
from openpyxl import load_workbook
from openpyxl.descriptors.base import String

wb = load_workbook(filename = 'names.xlsx')
ws = wb.active
names = []
names_no_dup = []
for row in ws.iter_rows(min_row=2, max_col=3):
   # names.append((row[0].value[::-1],row[1].value[::-1],row[2].value))
    names.append((row[0].value[::-1]))
s = list(set(names))
for n in s:
    names_no_dup.append([n])

for family in names_no_dup:
    for row in ws.iter_rows(min_row=2, max_col=3):
        if family[0] == row[0].value[::-1]:
            family.append([row[1].value[::-1],row[2].value])
print(names_no_dup)
